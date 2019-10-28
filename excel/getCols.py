# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active
cols = []
for row in ws.iter_cols():
	cols.append(row)

print(cols)  # 所有列
print(cols[0]) # 获取第一列
print(cols[0][0]) # 获取第一列第一列的单元格对象
print(cols[0][0].value)  # 获取第一列第一列的单元格对象的值

print(cols[len(cols) - 1])  # 获取最后列 print cols[-1]
print(cols[len(cols) - 1][len(cols[0]) - 1])  # 获取第后一行和最后一列的单元格对象
print(cols[len(cols) - 1][len(cols[0]) - 1].value)  # 获取第后一行和最后一列的单元格对象的值
