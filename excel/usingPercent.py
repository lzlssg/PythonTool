# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
#wb.guess_types = False
ws=wb.active
ws["D1"]="12%"
print (ws["D1"].value)
wb.save("sample.xlsx")
#结果会打印百分数
