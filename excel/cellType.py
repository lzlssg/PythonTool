 # -*- coding: utf-8 -*-
 from openpyxl import Workbook
 from openpyxl import load_workbook
 import datetime
 wb = load_workbook('sample.xlsx')

 ws=wb.active
 wb.guess_types = True

 ws["A1"]=datetime.datetime(2010, 7, 21)
 print (ws["A1"].number_format)

 ws["A2"]="12%"
 print (ws["A2"].number_format)

 ws["A3"]= 1.1
 print (ws["A4"].number_format)

 ws["A4"]= "中国"
 print (ws["A5"].number_format)
# Save the file
 wb.save("sample.xlsx")
