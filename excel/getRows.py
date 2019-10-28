# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws = wb.active
rows = []
for row in ws.iter_rows():
	rows.append(row)
print(rows)  # 所有行
print(rows[0]) # 获取第一行
print(rows[0][0]) # 获取第一行第一列的单元格对象
print(rows[0][0].value)  # 获取第一行第一列的单元格对象的值
print(rows[len(rows) - 1])  # 获取最后行 print rows[-1]
print(rows[len(rows) - 1][len(rows[0]) - 1])  # 获取第后一行和最后一列的单元格对象
print(rows[len(rows) - 1][len(rows[0]) - 1].value)  # 获取第后一行和最后一列的单元格对象的值
