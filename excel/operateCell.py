# -*- coding: utf-8 -*-

from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet("Mysheet")           #创建一个sheet

ws1["A1"]=123.11
ws1["B2"]="你好"
d = ws1.cell(row=4, column=2, value=10)

print (ws1["A1"].value)
print (ws1["B2"].value)
print (d.value)

# Save the file
wb.save("sample.xlsx")
