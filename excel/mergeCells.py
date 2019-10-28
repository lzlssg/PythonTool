# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
ws1=wb.active

ws1.merge_cells('A2:D2')
#.ws1.unmerge_cells('A2:D2')  #合并后的单元格，脚本单独执行拆分操作会报错，需要重新执行合并操作再拆分

# or equivalently
ws1.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
#ws1.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)

# Save the file
wb.save("sample.xlsx")
