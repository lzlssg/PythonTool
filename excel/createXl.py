# -*- coding: utf-8 -*-

from openpyxl import Workbook
wb = Workbook()    #创建文件对象

# grab the active worksheet
ws = wb.active     #获取第一个sheet

# Data can be assigned directly to cells
ws['A1'] = 42      #写入数字
ws['B1'] = "你好"+"automation test" #写入中文（unicode中文也可）

# Rows can also be appended
ws.append([1, 2, 3])    #写入多个单元格

# Python types will automatically be converted
import datetime
import time
ws['A2'] = datetime.datetime.now()    #写入一个当前时间
#写入一个自定义的时间格式
ws['A3'] =time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())

# Save the file
wb.save("sample.xlsx")
